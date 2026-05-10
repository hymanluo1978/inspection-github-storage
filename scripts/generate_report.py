#!/usr/bin/env python3
import os
import json
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
 
def load_json_files(data_dir):
records = []
for filename in os.listdir(data_dir):
if filename.endswith('.json'):
with open(os.path.join(data_dir, filename), 'r', encoding='utf-8') as f:
records.append(json.load(f))
return records
 
def create_overview_sheet(wb, records):
ws = wb.create_sheet('巡查总览', 0)
headers = ['记录ID', '项目名称', '楼号', '房号', '巡查时间', '提交时间']
ws.append(headers)
for record in records:
ws.append([
record.get('id', ''),
record.get('projectName', ''),
record.get('building', ''),
record.get('room', ''),
record.get('inspectionTime', ''),
record.get('createdAt', '')
])
style_header(ws, 1)
auto_adjust_width(ws)
 
def create_safety_sheet(wb, records):
ws = wb.create_sheet('安全文明巡查', 1)
headers = ['记录ID', '项目名称', '楼号', '房号', '巡查时间']
safety_items = []
if records and 'safety' in records[0]:
for item in records[0]['safety']:
headers.append(item['name'])
safety_items.append(item['name'])
ws.append(headers)
for record in records:
row = [
record.get('id', ''),
record.get('projectName', ''),
record.get('building', ''),
record.get('room', ''),
record.get('inspectionTime', '')
]
if 'safety' in record:
for item in record['safety']:
row.append(item.get('result', ''))
ws.append(row)
style_header(ws, 1)
highlight_problems(ws, safety_items)
auto_adjust_width(ws)
 
def create_progress_sheet(wb, records):
ws = wb.create_sheet('工序完成情况', 2)
headers = ['记录ID', '项目名称', '楼号', '房号', '巡查时间']
progress_items = []
if records and 'progress' in records[0]:
for item in records[0]['progress'].keys():
headers.append(item)
progress_items.append(item)
ws.append(headers)
for record in records:
row = [
record.get('id', ''),
record.get('projectName', ''),
record.get('building', ''),
record.get('room', ''),
record.get('inspectionTime', '')
]
if 'progress' in record:
for item in progress_items:
row.append(record['progress'].get(item, ''))
ws.append(row)
style_header(ws, 1)
auto_adjust_width(ws)
 
def create_inspection_sheet(wb, records, sheet_name, key, inspector_label='验收人', result_label='验收结论'):
ws = wb.create_sheet(sheet_name)
headers = ['记录ID', '项目名称', '楼号', '房号', inspector_label, result_label, '描述', '照片']
ws.append(headers)
for record in records:
if key in record.get('inspections', {}):
insp = record['inspections'][key]
photos = ', '.join(insp.get('photos', []))
ws.append([
record.get('id', ''),
record.get('projectName', ''),
record.get('building', ''),
record.get('room', ''),
insp.get('inspector', ''),
insp.get('result', ''),
insp.get('description', ''),
photos
])
style_header(ws, 1)
auto_adjust_width(ws)
 
def create_tile_sheet(wb, records):
ws = wb.create_sheet('墙地砖实测', 3)
headers = ['记录ID', '项目名称', '楼号', '房号', '实测人', '实测时间', 
'空鼓', '高低差', '垂直度', '平整度', '阴阳角']
ws.append(headers)
for record in records:
tile = record.get('tile', {})
ws.append([
record.get('id', ''),
record.get('projectName', ''),
record.get('building', ''),
record.get('room', ''),
tile.get('person', ''),
tile.get('time', ''),
tile.get('hollow', ''),
tile.get('height', ''),
tile.get('vertical', ''),
tile.get('flatness', ''),
tile.get('corner', '')
])
style_header(ws, 1)
auto_adjust_width(ws)
 
def create_putty_sheet(wb, records):
ws = wb.create_sheet('腻子面实测', 4)
headers = ['记录ID', '项目名称', '楼号', '房号', '实测人', '实测时间',
'顶棚极差', '阴阳角', '垂直度', '平整度', '三边两线', '阴阳角直线度']
ws.append(headers)
for record in records:
putty = record.get('putty', {})
ws.append([
record.get('id', ''),
record.get('projectName', ''),
record.get('building', ''),
record.get('room', ''),
putty.get('person', ''),
putty.get('time', ''),
putty.get('ceiling', ''),
putty.get('corner', ''),
putty.get('vertical', ''),
putty.get('flatness', ''),
putty.get('lines', ''),
putty.get('straightness', '')
])
style_header(ws, 1)
auto_adjust_width(ws)
 
def create_bathroom_sheet(wb, records):
ws = wb.create_sheet('卫生间蓄水验收', 5)
headers = ['记录ID', '项目名称', '楼号', '房号', '验收人', '验收结论', '照片']
ws.append(headers)
for record in records:
bath = record.get('bathroom', {})
photos = ', '.join(bath.get('photos', []))
ws.append([
record.get('id', ''),
record.get('projectName', ''),
record.get('building', ''),
record.get('room', ''),
bath.get('inspector', ''),
bath.get('result', ''),
photos
])
style_header(ws, 1)
auto_adjust_width(ws)
 
def style_header(ws, row):
header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
thin_border = Border(
left=Side(style='thin'),
right=Side(style='thin'),
top=Side(style='thin'),
bottom=Side(style='thin')
)
for cell in ws[row]:
cell.fill = header_fill
cell.font = header_font
cell.alignment = Alignment(horizontal='center', vertical='center')
cell.border = thin_border
 
def highlight_problems(ws, items):
warning_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
warning_font = Font(color='DC2626')
for row in range(2, ws.max_row + 1):
for col, item in enumerate(items, start=5):
cell = ws.cell(row=row, column=col)
if cell.value == '存在问题':
cell.fill = warning_fill
cell.font = warning_font
 
def auto_adjust_width(ws):
for column in ws.columns:
max_length = 0
column_letter = column[0].column_letter
for cell in column:
try:
if len(str(cell.value)) > max_length:
max_length = len(str(cell.value))
except:
pass
adjusted_width = min(max_length + 2, 50)
ws.column_dimensions[column_letter].width = adjusted_width
 
def main():
data_dir = 'data'
output_dir = 'reports'
if not os.path.exists(output_dir):
os.makedirs(output_dir)
if not os.path.exists(data_dir):
print('No data directory found')
return
records = load_json_files(data_dir)
if not records:
print('No records found')
return
wb = Workbook()
wb.remove(wb.active)
create_overview_sheet(wb, records)
create_safety_sheet(wb, records)
create_progress_sheet(wb, records)
create_tile_sheet(wb, records)
create_putty_sheet(wb, records)
if 'inspections' in records[0]:
create_inspection_sheet(wb, records, '轻钢龙骨验收', 'lightSteel')
create_inspection_sheet(wb, records, '防水切片验收', 'waterproof')
create_inspection_sheet(wb, records, '蓄水验收(厨卫)', 'waterStore')
create_inspection_sheet(wb, records, '墙布基层验收', 'wallBase')
create_bathroom_sheet(wb, records)
filename = os.path.join(output_dir, f'巡检验收台账_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx')
wb.save(filename)
print(f'Report generated: {filename}')
 
if __name__ == '__main__':
main()
